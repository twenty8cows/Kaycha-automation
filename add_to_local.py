import os
from openpyxl import load_workbook
from config.confidential_info import operating_file_path_instance


def find_files_with_suffix(folder_path, suffix):
    return [f for f in os.listdir(folder_path) if f.endswith(suffix)]


def load_excel_workbook(path, sheet_name=None, keep_vba=True):
    workbook = load_workbook(path, keep_vba=keep_vba)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    return workbook, worksheet


def find_first_empty_row(worksheet):
    for idx, row in enumerate(worksheet.iter_rows(min_col=1, max_col=1), start=1):
        if row[0].value is None:
            return idx
    return None


def append_data_to_worksheet(source_sheet, destination_worksheet, start_row):
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        for col_index, cell_value in enumerate(row, start=1):
            destination_worksheet.cell(
                row=start_row, column=col_index, value=cell_value)
        start_row += 1


def main():
    operating_file_path = operating_file_path_instance.operating_file_path
    excel_filename = operating_file_path_instance.excel_filename
    full_excel_path = os.path.join(operating_file_path, excel_filename)

    workbook, worksheet = load_excel_workbook(full_excel_path, "Kaycha")
    first_empty_row = find_first_empty_row(worksheet)

    downloads_folder_path = os.path.expanduser('~/Downloads')
    processed_files = find_files_with_suffix(
        downloads_folder_path, '_processed.xlsx')

    if not processed_files:
        print("No processed file found")
        exit()

    processed_file_path = os.path.join(
        downloads_folder_path, processed_files[0])
    print(f"Found processed file: {processed_file_path}")

    # Here, if sheet_name is None, the function will return the active sheet
    processed_workbook, processed_sheet = load_excel_workbook(
        processed_file_path)
    if not processed_sheet:  # If no specific sheet name is given, use the active sheet
        processed_sheet = processed_workbook.active

    append_data_to_worksheet(processed_sheet, worksheet, first_empty_row)

    workbook.save(full_excel_path)
    print(f"Data appended successfully to {full_excel_path}")


if __name__ == "__main__":
    main()
